from datetime import date
from io import BytesIO

from django.conf import settings
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.core.validators import validate_email
from django.http import HttpResponse
from django.utils import timezone
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from rest_framework import permissions, status, viewsets, generics
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import Alerta, Categoria, Producto, Movimiento, Pedido, UserProfile
from .serializers import (
    AlertaSerializer,
    CategoriaSerializer,
    ProductoSerializer,
    MovimientoSerializer,
    PedidoSerializer,
    UserManagementSerializer,
)


class IsStaffOrSuperuser(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser))


class IsInventoryManagerOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.is_staff or request.user.is_superuser


def _is_email_domain_reachable(domain: str) -> bool:
    domain = (domain or '').strip().lower()
    if not domain:
        return False

    # Valida resolución DNS básica del dominio de correo.
    import socket

    try:
        socket.getaddrinfo(domain, None)
        return True
    except socket.gaierror:
        return False


def _validate_registration_email(email: str):
    normalized_email = (email or '').strip().lower()
    if not normalized_email:
        raise ValidationError('El correo es obligatorio.')

    try:
        validate_email(normalized_email)
    except ValidationError:
        raise ValidationError('El formato del correo no es válido.')

    domain = normalized_email.split('@')[-1]
    blocked_domains = {'example.com', 'test.com', 'mailinator.com', 'invalid.com'}
    if domain in blocked_domains:
        raise ValidationError('El dominio del correo no está permitido.')

    if not _is_email_domain_reachable(domain):
        raise ValidationError('No se pudo validar el dominio del correo. Verifica que exista.')

    return normalized_email


def _get_role_from_user(user: User) -> str:
    return 'jefe' if user.is_superuser else 'admin' if user.is_staff else 'usuario'


def _build_verification_link(request, user: User) -> str:
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)
    frontend_base = getattr(settings, 'FRONTEND_APP_URL', '').rstrip('/')
    if frontend_base:
        return f'{frontend_base}/verify-email/{uid}/{token}'

    origin = request.build_absolute_uri('/').rstrip('/')
    return f'{origin}/verify-email/{uid}/{token}'


def _send_verification_email(request, user: User) -> None:
    verification_link = _build_verification_link(request, user)
    subject = 'SIGIRL - Verificación de correo'
    message = (
        'Hola,\n\n'
        'Para activar tu cuenta en SIGIRL debes verificar tu correo.\n'
        f'Enlace de verificación: {verification_link}\n\n'
        'Si no solicitaste este registro, ignora este mensaje.'
    )

    send_mail(
        subject,
        message,
        getattr(settings, 'DEFAULT_FROM_EMAIL', 'no-reply@sigirl.local'),
        [user.email],
        fail_silently=False,
    )

    profile, _ = UserProfile.objects.get_or_create(user=user)
    profile.email_verification_sent_at = timezone.now()
    profile.save(update_fields=['email_verification_sent_at'])


# =====================
# TOKEN (PUBLIC)
# =====================
class PublicTokenObtainPairView(TokenObtainPairView):
    permission_classes = [AllowAny]

    # Sobrescribe el login JWT para devolver también el rol del usuario.
    def post(self, request, *args, **kwargs):
        username = (request.data.get('username') or '').strip()
        password = request.data.get('password') or ''
        user = User.objects.filter(username=username).first()

        if user and user.check_password(password):
            profile, _ = UserProfile.objects.get_or_create(user=user)
            if not profile.email_verified:
                return Response(
                    {
                        'error': 'Tu correo no está verificado. Revisa tu bandeja y confirma tu cuenta.',
                        'email_not_verified': True,
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )

        response = super().post(request, *args, **kwargs)

        if response.status_code == 200:
            try:
                user = User.objects.get(username=username)
                role = _get_role_from_user(user)

                response.data["role"] = role
                response.data["username"] = user.username
            except User.DoesNotExist:
                pass

        return response


class PublicTokenRefreshView(TokenRefreshView):
    permission_classes = [AllowAny]


# =====================
# PRODUCTO (VIEWSET)
# =====================
class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    permission_classes = [IsInventoryManagerOrReadOnly]


# =====================
# CATEGORIA
# =====================
class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [IsAuthenticated]


# =====================
# MOVIMIENTO
# =====================
class MovimientoViewSet(viewsets.ModelViewSet):
    queryset = Movimiento.objects.all()
    serializer_class = MovimientoSerializer
    permission_classes = [IsAuthenticated]


# =====================
# PEDIDO
# =====================
class PedidoViewSet(viewsets.ModelViewSet):
    queryset = Pedido.objects.all()
    serializer_class = PedidoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset().select_related('producto', 'usuario')
        if user.is_staff or user.is_superuser:
            return queryset
        return queryset.filter(usuario=user)

    def perform_create(self, serializer):
        user = self.request.user
        profile = getattr(user, 'profile', None)
        serializer.save(
            usuario=user,
            solicitante=user.get_full_name().strip() or user.username,
            creado_por=user.username,
            departamento=getattr(profile, 'department', '') or '',
        )

    def update(self, request, *args, **kwargs):
        pedido = self.get_object()
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({'error': 'No tienes permiso para actualizar pedidos.'}, status=status.HTTP_403_FORBIDDEN)

        payload = request.data.copy()
        estado = str(payload.get('estado', pedido.estado)).lower()
        estado_anterior = pedido.estado

        if estado == 'rechazado' and not str(payload.get('motivo_rechazo', pedido.motivo_rechazo or '')).strip():
            return Response({'motivo_rechazo': ['Debes indicar un motivo de rechazo.']}, status=status.HTTP_400_BAD_REQUEST)

        if estado == 'aprobado' and estado_anterior != 'aprobado':
            producto = pedido.producto
            if producto.cantidad >= pedido.cantidad:
                producto.cantidad -= pedido.cantidad
                producto.save()
            else:
                return Response({'error': 'No hay suficiente stock'}, status=status.HTTP_400_BAD_REQUEST)

        payload['estado'] = estado
        if estado in ('aprobado', 'rechazado'):
            payload['fecha_respuesta'] = date.today()

        serializer = self.get_serializer(pedido, data=payload, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(serializer.data)


# =====================
# REGISTER
# =====================
@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    username = (request.data.get('username') or '').strip()
    email = (request.data.get('email') or '').strip()
    password = request.data.get('password') or ''
    first_name = (request.data.get('first_name') or '').strip()
    last_name = (request.data.get('last_name') or '').strip()
    requested_role = request.data.get('role', 'usuario')
    institution = (request.data.get('institution') or '').strip()
    department = (request.data.get('department') or '').strip()

    if not username or not password or not email:
        return Response({'error': 'Debes enviar usuario, correo y contraseña.'}, status=status.HTTP_400_BAD_REQUEST)

    if User.objects.filter(username__iexact=username).exists():
        return Response({'username': ['El usuario ya existe.']}, status=status.HTTP_400_BAD_REQUEST)

    if User.objects.filter(email__iexact=email).exists():
        return Response({'email': ['Ese correo ya está registrado.']}, status=status.HTTP_400_BAD_REQUEST)

    try:
        email = _validate_registration_email(email)
    except ValidationError as exc:
        return Response({'email': [str(exc)]}, status=status.HTTP_400_BAD_REQUEST)

    user = User.objects.create_user(
        username=username,
        email=email,
        password=password,
        first_name=first_name,
        last_name=last_name,
    )

    if requested_role == 'admin':
        user.is_staff = True
    elif requested_role in ['jefe', 'jefe_superior']:
        user.is_staff = True
        user.is_superuser = True

    user.save()

    profile, _ = UserProfile.objects.get_or_create(user=user)
    profile.institution = institution
    profile.department = department
    profile.email_verified = False
    profile.email_verified_at = None
    profile.save()

    try:
        _send_verification_email(request, user)
    except Exception:
        user.delete()
        return Response(
            {'error': 'No se pudo enviar el correo de verificación. Intenta de nuevo en unos minutos.'},
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    return Response(
        {
            'mensaje': 'Usuario creado. Revisa tu correo para activar la cuenta.',
            'requires_verification': True,
            'username': user.username,
            'email': user.email,
            'role': _get_role_from_user(user),
        },
        status=status.HTTP_201_CREATED,
    )


@api_view(['GET'])
@permission_classes([AllowAny])
def verify_email(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except Exception:
        return Response({'error': 'El enlace de verificación no es válido.'}, status=status.HTTP_400_BAD_REQUEST)

    if not default_token_generator.check_token(user, token):
        return Response({'error': 'El enlace de verificación expiró o ya no es válido.'}, status=status.HTTP_400_BAD_REQUEST)

    profile, _ = UserProfile.objects.get_or_create(user=user)
    profile.email_verified = True
    profile.email_verified_at = timezone.now()
    profile.save(update_fields=['email_verified', 'email_verified_at'])

    return Response({'mensaje': 'Correo verificado correctamente. Ya puedes iniciar sesión.'}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([AllowAny])
def resend_verification_email(request):
    username = (request.data.get('username') or '').strip()
    email = (request.data.get('email') or '').strip().lower()

    user = None
    if username:
        user = User.objects.filter(username__iexact=username).first()
    elif email:
        user = User.objects.filter(email__iexact=email).first()

    if not user:
        return Response({'error': 'No encontramos una cuenta con esos datos.'}, status=status.HTTP_404_NOT_FOUND)

    profile, _ = UserProfile.objects.get_or_create(user=user)
    if profile.email_verified:
        return Response({'mensaje': 'La cuenta ya está verificada.'}, status=status.HTTP_200_OK)

    try:
        _send_verification_email(request, user)
    except Exception:
        return Response({'error': 'No fue posible reenviar el correo ahora. Intenta nuevamente.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

    return Response({'mensaje': 'Se reenviaron las instrucciones de verificación a tu correo.'}, status=status.HTTP_200_OK)


# =====================
# AUTH - Get Current User
# =====================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_current_user(request):
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)
    role = _get_role_from_user(user)
    return Response({
        "id": user.id,
        "username": user.username,
        "email": user.email,
        "email_verified": profile.email_verified,
        "is_staff": user.is_staff,
        "is_superuser": user.is_superuser,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "role": role,
    })


# =====================
# AUTH - Manage Profile
# =====================
def _serialize_profile_response(user, profile):
    role = _get_role_from_user(user)
    return {
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'email_verified': profile.email_verified,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'role': role,
        'profile': {
            'institution': profile.institution,
            'department': profile.department,
            'phone': profile.phone,
            'cargo': profile.cargo,
            'bio': profile.bio,
            'avatar': profile.avatar,
        }
    }


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def manage_profile(request):
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)

    if request.method == 'GET':
        return Response(_serialize_profile_response(user, profile))

    if request.method == 'DELETE':
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    data = request.data
    username = str(data.get('username', user.username)).strip()
    email = str(data.get('email', user.email)).strip()

    if username and username != user.username and User.objects.exclude(pk=user.pk).filter(username=username).exists():
        return Response({'username': ['Ya existe un usuario con ese nombre.']}, status=status.HTTP_400_BAD_REQUEST)
    if email and email != user.email and User.objects.exclude(pk=user.pk).filter(email=email).exists():
        return Response({'email': ['Ya existe un usuario con ese correo.']}, status=status.HTTP_400_BAD_REQUEST)

    user.username = username or user.username
    user.first_name = data.get('first_name', user.first_name)
    user.last_name = data.get('last_name', user.last_name)
    user.email = email
    user.save()

    profile_data = data.get('profile', {}) if isinstance(data.get('profile', {}), dict) else {}
    profile.institution = profile_data.get('institution', data.get('institution', profile.institution))
    profile.department = profile_data.get('department', data.get('department', profile.department))
    profile.phone = profile_data.get('phone', data.get('phone', profile.phone))
    profile.cargo = profile_data.get('cargo', data.get('cargo', profile.cargo))
    profile.bio = profile_data.get('bio', data.get('bio', profile.bio))
    profile.avatar = profile_data.get('avatar', data.get('avatar', profile.avatar))
    profile.save()

    return Response(_serialize_profile_response(user, profile))


@api_view(['GET'])
@permission_classes([AllowAny])
def download_inventory_template_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Plantilla Inventario'

    headers = ['nombre', 'cantidad', 'categoria', 'fecha']
    sheet.append(headers)
    sheet.append(['Acetona', 25, 'Solventes', '2026-04-27'])
    sheet.append(['Guantes Nitrilo', 100, 'EPP', '2026-04-27'])

    header_fill = PatternFill(start_color='1FA971', end_color='1FA971', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    centered = Alignment(horizontal='center', vertical='center')

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered
        sheet.column_dimensions[chr(64 + column_index)].width = max(14, len(header) + 5)

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_inventario_sigirl.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([AllowAny])
def download_inventory_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Inventario'

    headers = ['Nombre', 'Tipo', 'Categoria', 'Cantidad', 'Umbral minimo', 'Ubicacion', 'Estado']
    sheet.append(headers)

    productos = Producto.objects.select_related('categoria').order_by('nombre')
    for producto in productos:
        sheet.append([
            producto.nombre,
            producto.tipo,
            producto.categoria.nombre if producto.categoria else '',
            producto.cantidad,
            producto.minimo,
            producto.ubicacion or '',
            producto.estado,
        ])

    header_fill = PatternFill(start_color='0F7A53', end_color='0F7A53', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for column_index in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions[chr(64 + column_index)].width = 20

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="inventario_sigirl.xlsx"'
    return response


def _register_pdf_font():
    import importlib

    pdfmetrics = importlib.import_module('reportlab.pdfbase.pdfmetrics')
    ttfonts = importlib.import_module('reportlab.pdfbase.ttfonts')
    TTFont = getattr(ttfonts, 'TTFont')

    candidate_fonts = [
        ('SegoeUI', 'C:/Windows/Fonts/segoeui.ttf'),
        ('Calibri', 'C:/Windows/Fonts/calibri.ttf'),
    ]
    for font_name, font_path in candidate_fonts:
        try:
            pdfmetrics.registerFont(TTFont(font_name, font_path))
            return font_name
        except Exception:
            continue
    return 'Helvetica'


@api_view(['GET'])
@permission_classes([AllowAny])
def download_inventory_pdf(request):
    import importlib

    colors = importlib.import_module('reportlab.lib.colors')
    A4 = getattr(importlib.import_module('reportlab.lib.pagesizes'), 'A4')
    ParagraphStyle = getattr(importlib.import_module('reportlab.lib.styles'), 'ParagraphStyle')
    mm = getattr(importlib.import_module('reportlab.lib.units'), 'mm')
    platypus = importlib.import_module('reportlab.platypus')
    Paragraph = getattr(platypus, 'Paragraph')
    SimpleDocTemplate = getattr(platypus, 'SimpleDocTemplate')
    Spacer = getattr(platypus, 'Spacer')
    Table = getattr(platypus, 'Table')
    TableStyle = getattr(platypus, 'TableStyle')

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=18 * mm,
        bottomMargin=14 * mm,
    )

    font_name = _register_pdf_font()
    title_style = ParagraphStyle(
        name='Title',
        fontName=font_name,
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#0F7A53'),
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        name='Subtitle',
        fontName=font_name,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#4B5563'),
        spaceAfter=12,
    )

    productos = Producto.objects.select_related('categoria').order_by('nombre')
    table_data = [['Producto', 'Categoria', 'Cantidad', 'Minimo', 'Ubicacion', 'Estado']]
    for p in productos:
        table_data.append([
            p.nombre,
            p.categoria.nombre if p.categoria else '',
            str(p.cantidad),
            str(p.minimo),
            p.ubicacion or '-',
            p.estado,
        ])

    table = Table(table_data, repeatRows=1, colWidths=[70 * mm, 34 * mm, 20 * mm, 20 * mm, 30 * mm, 22 * mm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8.5),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1FA971')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (2, 1), (3, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#F3FBF7')]),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#C6E8D9')),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
        ('TOPPADDING', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
    ]))

    story = [
        Paragraph('SIGIRL - Reporte de Inventario', title_style),
        Paragraph(f'Generado: {timezone.localtime().strftime("%Y-%m-%d %H:%M")}', subtitle_style),
        Spacer(1, 4),
        table,
    ]
    document.build(story)

    pdf_data = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_data, content_type='application/pdf; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario_sigirl.pdf"'
    return response


# =====================
# ALERTA (VIEWSET)
# =====================
class AlertaViewSet(viewsets.ModelViewSet):
    queryset = Alerta.objects.all().order_by('-fecha')
    serializer_class = AlertaSerializer
    permission_classes = [IsAuthenticated]


# =====================
# USER MANAGEMENT (VIEWSET)
# =====================
class UserManagementViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().order_by('username')
    serializer_class = UserManagementSerializer
    permission_classes = [IsStaffOrSuperuser]

    def create(self, request, *args, **kwargs):
        data = request.data
        username = str(data.get('username') or data.get('nombre_input') or '').strip()
        email = str(data.get('email') or '').strip()
        password = str(data.get('password') or '').strip()

        if not username:
            return Response({'username': ['El nombre de usuario es obligatorio.']}, status=status.HTTP_400_BAD_REQUEST)
        if not password:
            return Response({'password': ['La contraseña es obligatoria.']}, status=status.HTTP_400_BAD_REQUEST)
        if User.objects.filter(username=username).exists():
            return Response({'username': ['Ya existe un usuario con ese nombre.']}, status=status.HTTP_400_BAD_REQUEST)
        if email and User.objects.filter(email=email).exists():
            return Response({'email': ['Ya existe un usuario con ese correo.']}, status=status.HTTP_400_BAD_REQUEST)

        user = User(username=username, email=email)
        nombre = str(data.get('nombre_completo') or data.get('full_name') or username).strip()
        parts = nombre.split(' ', 1)
        user.first_name = parts[0]
        user.last_name = parts[1] if len(parts) > 1 else ''

        rol = data.get('rol_input') or data.get('rol') or 'usuario'
        if rol == 'jefe':
            user.is_staff = True
            user.is_superuser = True
        elif rol == 'admin':
            user.is_staff = True
            user.is_superuser = False
        else:
            user.is_staff = False
            user.is_superuser = False

        user.set_password(password)
        user.save()

        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.department = str(data.get('departamento_input') or data.get('department') or '').strip()
        profile.save()

        serializer = self.get_serializer(user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        user = self.get_object()
        data = request.data

        username = str(data.get('username') or '').strip()
        if username and username != user.username:
            if User.objects.exclude(pk=user.pk).filter(username=username).exists():
                return Response({'username': ['Ya existe un usuario con ese nombre.']}, status=status.HTTP_400_BAD_REQUEST)
            user.username = username

        nombre = str(data.get('nombre_input') or data.get('nombre_completo') or '').strip()
        if nombre:
            parts = nombre.split(' ', 1)
            user.first_name = parts[0]
            user.last_name = parts[1] if len(parts) > 1 else ''

        email = str(data.get('email') or '').strip()
        if email:
            if email != user.email and User.objects.exclude(pk=user.pk).filter(email=email).exists():
                return Response({'email': ['Ya existe un usuario con ese correo.']}, status=status.HTTP_400_BAD_REQUEST)
            user.email = email

        rol = data.get('rol_input') or data.get('rol') or ''
        if rol == 'jefe':
            user.is_staff = True
            user.is_superuser = True
        elif rol == 'admin':
            user.is_staff = True
            user.is_superuser = False
        elif rol == 'usuario':
            user.is_staff = False
            user.is_superuser = False

        password = data.get('password', '')
        if password:
            user.set_password(password)

        user.save()

        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.department = str(data.get('departamento_input') or data.get('department') or '').strip()
        profile.save()

        serializer = self.get_serializer(user)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        if user == request.user:
            return Response({"error": "No puedes eliminar tu propia cuenta"}, status=400)
        user.delete()
        return Response(status=204)
